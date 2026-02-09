from __future__ import annotations

import getpass
import json
import logging
import os
import platform
import uuid
from datetime import datetime, timedelta
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger("auditoria_excel")

DEFAULT_PATH = r"P:\AuditoriaRobos\Auditoria_Robos.xlsx"
PENDING_DIR = r"P:\AuditoriaRobos\pending"

RUNS_SHEET = "RUNS"
ERRORS_SHEET = "ERRORS"
ROBOS_SHEET = "ROBOS"
DASHBOARD_SHEET = "DASHBOARD"

RUNS_COLUMNS = [
    "run_id",
    "robo_nome",
    "versao_robo",
    "ambiente",
    "data_execucao",
    "hora_inicio",
    "hora_fim",
    "duracao_segundos",
    "total_processado",
    "total_sucesso",
    "total_erro",
    "taxa_sucesso",
    "taxa_erro",
    "erros_auto_mitigados",
    "erros_manuais",
    "erros_pendentes",
    "resultado_final",
    "observacoes",
    "host_maquina",
    "usuario_execucao",
    "origem_codigo",
    "commit_hash",
    "build_id",
]

ERRORS_COLUMNS = [
    "run_id",
    "robo_nome",
    "timestamp",
    "etapa",
    "tipo_erro",
    "codigo_erro",
    "mensagem_resumida",
    "registro_id",
    "mitigacao",
    "resolvido_em",
]

ROBOS_COLUMNS = [
    "robo_nome",
    "versao_robo",
    "origem_codigo",
    "ambiente",
    "ativo",
    "ultima_execucao",
    "owner",
    "observacoes",
]

ENV_OPTIONS = ["PROD", "TEST"]
RESULTADO_OPTIONS = ["100% Concluído", "Concluído com Exceções", "Incompleto"]
TIPO_ERRO_OPTIONS = ["Tecnico", "Regra de Negocio", "Dados Invalidos", "Externo/Indisponibilidade"]
MITIGACAO_OPTIONS = ["Auto", "Manual", "Pendente"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def log_run(
    run_data: Dict,
    errors: Optional[List[Dict]] = None,
    path: str = DEFAULT_PATH,
) -> str:
    run_data = dict(run_data or {})
    errors = list(errors or [])

    run_id = run_data.get("run_id") or str(uuid.uuid4())
    run_data["run_id"] = run_id
    run_data.setdefault("robo_nome", "ASO")
    run_data.setdefault("origem_codigo", r"P:\ProcessosASO\codigos")
    run_data.setdefault("usuario_execucao", getpass.getuser())
    run_data.setdefault("host_maquina", platform.node())
    run_data.setdefault("versao_robo", get_project_version())
    run_data.setdefault("ambiente", os.getenv("ASO_ENV", "PROD"))

    started_at = _parse_dt(run_data.get("started_at"))
    finished_at = _parse_dt(run_data.get("finished_at"))
    if finished_at is None:
        finished_at = datetime.now()
    if started_at is None:
        duration = _safe_int(run_data.get("duracao_segundos"))
        if duration is not None:
            started_at = finished_at - timedelta(seconds=duration)
        else:
            started_at = finished_at

    if run_data.get("duracao_segundos") in (None, ""):
        run_data["duracao_segundos"] = max(int((finished_at - started_at).total_seconds()), 0)

    run_data.setdefault("data_execucao", started_at.date())
    run_data.setdefault("hora_inicio", started_at.time())
    run_data.setdefault("hora_fim", finished_at.time())

    run_data.setdefault("total_processado", 0)
    run_data.setdefault("total_sucesso", 0)
    run_data.setdefault("total_erro", 0)
    run_data.setdefault("erros_auto_mitigados", 0)
    run_data.setdefault("erros_manuais", 0)
    run_data.setdefault("observacoes", "")
    run_data.setdefault("commit_hash", _try_get_git_commit())
    run_data.setdefault("build_id", "")

    if not run_data.get("resultado_final"):
        run_data["resultado_final"] = _calc_resultado_final(
            run_data.get("total_processado"),
            run_data.get("total_sucesso"),
            run_data.get("total_erro"),
            run_data.get("erros_auto_mitigados"),
            run_data.get("erros_manuais"),
        )

    target_dir = os.path.dirname(path)
    if target_dir:
        os.makedirs(target_dir, exist_ok=True)

    wb, created, invalid_existing = _ensure_workbook(path)
    ws_runs = wb[RUNS_SHEET]
    ws_errors = wb[ERRORS_SHEET]
    ws_robos = wb[ROBOS_SHEET]

    _append_run(ws_runs, run_data)
    _append_errors(ws_errors, run_data, errors)
    _upsert_robo(ws_robos, run_data)

    if invalid_existing:
        logger.warning("Arquivo de auditoria estava invalido. Um novo XLSX foi criado.")

    try:
        wb.save(path)
        logger.info("Auditoria Excel salva em %s", path)
    except PermissionError:
        _save_pending(wb, run_data, errors)
    except OSError:
        _save_pending(wb, run_data, errors)

    return run_id


def _ensure_workbook(path: str) -> Tuple[Workbook, bool, bool]:
    created = False
    invalid_existing = False
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
        except Exception as exc:
            logger.error("Falha ao abrir XLSX existente (%s): %s", path, exc)
            _backup_corrupt_file(path)
            wb = Workbook()
            created = True
            invalid_existing = True
    else:
        wb = Workbook()
        created = True

    if created:
        ws = wb.active
        ws.title = RUNS_SHEET
        _setup_runs_sheet(ws)
        _setup_errors_sheet(wb.create_sheet(ERRORS_SHEET))
        _setup_robos_sheet(wb.create_sheet(ROBOS_SHEET))
        _setup_dashboard_sheet(wb.create_sheet(DASHBOARD_SHEET))
    else:
        if RUNS_SHEET not in wb.sheetnames:
            _setup_runs_sheet(wb.create_sheet(RUNS_SHEET, 0))
        if ERRORS_SHEET not in wb.sheetnames:
            _setup_errors_sheet(wb.create_sheet(ERRORS_SHEET))
        if ROBOS_SHEET not in wb.sheetnames:
            _setup_robos_sheet(wb.create_sheet(ROBOS_SHEET))
        if DASHBOARD_SHEET not in wb.sheetnames:
            _setup_dashboard_sheet(wb.create_sheet(DASHBOARD_SHEET))
    return wb, created, invalid_existing


def _backup_corrupt_file(path: str) -> None:
    try:
        base_dir = os.path.dirname(path)
        name, ext = os.path.splitext(os.path.basename(path))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(base_dir, f"{name}__CORROMPIDO__{timestamp}{ext or '.xlsx'}")
        if os.path.exists(path):
            os.replace(path, backup_path)
            logger.info("Backup do arquivo corrompido criado em %s", backup_path)
    except Exception as exc:
        logger.error("Falha ao mover arquivo corrompido: %s", exc)


def _setup_runs_sheet(ws) -> None:
    ws.append(RUNS_COLUMNS)
    _style_header(ws, len(RUNS_COLUMNS))
    _set_column_widths(ws, {
        "A": 36,
        "B": 12,
        "C": 12,
        "D": 10,
        "E": 12,
        "F": 10,
        "G": 10,
        "H": 14,
        "I": 14,
        "J": 12,
        "K": 10,
        "L": 12,
        "M": 10,
        "N": 16,
        "O": 12,
        "P": 14,
        "Q": 22,
        "R": 30,
        "S": 18,
        "T": 18,
        "U": 28,
        "V": 14,
        "W": 14,
    })
    ws.freeze_panes = "A2"
    _ensure_runs_table(ws)
    _add_runs_validations(ws)
    _add_runs_conditional_formats(ws)


def _setup_errors_sheet(ws) -> None:
    ws.append(ERRORS_COLUMNS)
    _style_header(ws, len(ERRORS_COLUMNS))
    _set_column_widths(ws, {
        "A": 36,
        "B": 12,
        "C": 20,
        "D": 20,
        "E": 22,
        "F": 14,
        "G": 60,
        "H": 20,
        "I": 12,
        "J": 18,
    })
    ws.freeze_panes = "A2"
    _ensure_errors_table(ws)
    _add_errors_validations(ws)


def _setup_robos_sheet(ws) -> None:
    ws.append(ROBOS_COLUMNS)
    _style_header(ws, len(ROBOS_COLUMNS))
    _set_column_widths(ws, {
        "A": 16,
        "B": 12,
        "C": 28,
        "D": 10,
        "E": 10,
        "F": 20,
        "G": 18,
        "H": 30,
    })
    ws.freeze_panes = "A2"
    _ensure_robos_table(ws)


def _setup_dashboard_sheet(ws) -> None:
    ws["A1"] = "Auditoria de Robos - Dashboard"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="left")

    kpi_labels = [
        ("A3", "Execucoes 7 dias"),
        ("A4", "Execucoes 30 dias"),
        ("A5", "Taxa media de sucesso (30d)"),
        ("A6", "Total erros (30d)"),
        ("A7", "% mitigado auto (30d)"),
        ("A8", "Tempo medio execucao (30d)"),
    ]
    for cell, label in kpi_labels:
        ws[cell] = label
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill("solid", fgColor="2F75B5")
        ws[cell].alignment = Alignment(horizontal="left")
        ws[cell].border = CELL_BORDER

    ws["B3"] = '=COUNTIFS(tbl_runs[data_execucao],">="&TODAY()-6)'
    ws["B4"] = '=COUNTIFS(tbl_runs[data_execucao],">="&TODAY()-29)'
    ws["B5"] = '=IFERROR(AVERAGEIFS(tbl_runs[taxa_sucesso],tbl_runs[data_execucao],">="&TODAY()-29),0)'
    ws["B6"] = '=IFERROR(SUMIFS(tbl_runs[total_erro],tbl_runs[data_execucao],">="&TODAY()-29),0)'
    ws["B7"] = '=IFERROR(SUMIFS(tbl_runs[erros_auto_mitigados],tbl_runs[data_execucao],">="&TODAY()-29)/SUMIFS(tbl_runs[total_erro],tbl_runs[data_execucao],">="&TODAY()-29),0)'
    ws["B8"] = '=IFERROR(AVERAGEIFS(tbl_runs[duracao_segundos],tbl_runs[data_execucao],">="&TODAY()-29),0)'

    for cell in ("B3", "B4", "B6", "B8"):
        ws[cell].font = Font(bold=True)
        ws[cell].border = CELL_BORDER
    ws["B5"].number_format = "0.00%"
    ws["B7"].number_format = "0.00%"
    ws["B5"].font = Font(bold=True)
    ws["B7"].font = Font(bold=True)
    ws["B5"].border = CELL_BORDER
    ws["B7"].border = CELL_BORDER

    ws["D2"] = "Taxa de Sucesso ao Longo do Tempo"
    ws["D2"].font = Font(bold=True)
    _add_success_line_chart(ws, "D3")

    ws["D18"] = "Total Processado vs Erros"
    ws["D18"].font = Font(bold=True)
    _add_volume_bar_chart(ws, "D19")

    ws["L2"] = "Tipos de Erro"
    ws["L2"].font = Font(bold=True)
    _add_error_pie_chart(ws, "L3")

    ws["L18"] = "Duracao por Execucao"
    ws["L18"].font = Font(bold=True)
    _add_duration_line_chart(ws, "L19")


def _append_run(ws, run_data: Dict) -> None:
    row_idx = _next_data_row(ws)
    for col_idx, col_name in enumerate(RUNS_COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if col_name == "taxa_sucesso":
            cell.value = f"=IF($I{row_idx}=0,0,$J{row_idx}/$I{row_idx})"
            cell.number_format = "0.00%"
        elif col_name == "taxa_erro":
            cell.value = f"=IF($I{row_idx}=0,0,$K{row_idx}/$I{row_idx})"
            cell.number_format = "0.00%"
        elif col_name == "erros_pendentes":
            cell.value = f"=MAX($K{row_idx}-$N{row_idx}-$O{row_idx},0)"
        else:
            cell.value = run_data.get(col_name)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(vertical="center")

    _format_runs_row(ws, row_idx)
    _ensure_runs_table(ws)


def _append_errors(ws, run_data: Dict, errors: List[Dict]) -> None:
    if not errors:
        return

    for err in errors:
        row_idx = _next_data_row(ws)
        for col_idx, col_name in enumerate(ERRORS_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == "run_id":
                cell.value = err.get("run_id") or run_data.get("run_id")
            elif col_name == "robo_nome":
                cell.value = err.get("robo_nome") or run_data.get("robo_nome")
            elif col_name == "timestamp":
                cell.value = _parse_dt(err.get("timestamp")) or datetime.now()
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            else:
                cell.value = err.get(col_name)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center")

    _ensure_errors_table(ws)


def _upsert_robo(ws, run_data: Dict) -> None:
    robo_nome = run_data.get("robo_nome")
    if not robo_nome:
        return

    row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == robo_nome:
            row = r
            break

    if row is None:
        row = _next_data_row(ws)

    values = {
        "robo_nome": robo_nome,
        "versao_robo": run_data.get("versao_robo"),
        "origem_codigo": run_data.get("origem_codigo"),
        "ambiente": run_data.get("ambiente"),
        "ativo": "Sim",
        "ultima_execucao": _parse_dt(run_data.get("finished_at")) or datetime.now(),
        "owner": run_data.get("owner", ""),
        "observacoes": run_data.get("observacoes", ""),
    }

    for col_idx, col_name in enumerate(ROBOS_COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = values.get(col_name)
        if col_name == "ultima_execucao":
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
        cell.border = CELL_BORDER

    _ensure_robos_table(ws)


def _ensure_runs_table(ws) -> None:
    _ensure_table(ws, "tbl_runs", len(RUNS_COLUMNS))


def _ensure_errors_table(ws) -> None:
    _ensure_table(ws, "tbl_errors", len(ERRORS_COLUMNS))


def _ensure_robos_table(ws) -> None:
    _ensure_table(ws, "tbl_robos", len(ROBOS_COLUMNS))


def _ensure_table(ws, name: str, total_cols: int) -> None:
    last_row = max(ws.max_row, 2)
    last_col = get_column_letter(total_cols)
    ref = f"A1:{last_col}{last_row}"
    if name in ws.tables:
        ws.tables[name].ref = ref
        return

    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def _add_runs_validations(ws) -> None:
    _add_list_validation(ws, "D2:D10000", ENV_OPTIONS)
    _add_list_validation(ws, "Q2:Q10000", RESULTADO_OPTIONS)


def _add_errors_validations(ws) -> None:
    _add_list_validation(ws, "E2:E10000", TIPO_ERRO_OPTIONS)
    _add_list_validation(ws, "I2:I10000", MITIGACAO_OPTIONS)


def _add_list_validation(ws, cell_range: str, values: Iterable[str]) -> None:
    formula = '"' + ",".join(values) + '"'
    for dv in ws.data_validations.dataValidation:
        if dv.formula1 == formula and dv.type == "list":
            return
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _add_runs_conditional_formats(ws) -> None:
    mismatch_formula = "=$J2+$K2<>$I2"
    mismatch_rule = FormulaRule(formula=[mismatch_formula], fill=PatternFill("solid", fgColor="FFC7CE"))
    ws.conditional_formatting.add("A2:W10000", mismatch_rule)

    error_rule = CellIsRule(operator="greaterThan", formula=["0.1"], fill=PatternFill("solid", fgColor="FFC7CE"))
    ws.conditional_formatting.add("M2:M10000", error_rule)

    success_rule = FormulaRule(formula=['=$Q2="100% Concluído"'], fill=PatternFill("solid", fgColor="C6EFCE"))
    ws.conditional_formatting.add("Q2:Q10000", success_rule)


def _style_header(ws, total_cols: int) -> None:
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER


def _set_column_widths(ws, widths: Dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _next_data_row(ws) -> int:
    if ws.max_row == 1:
        return 2
    if ws.max_row == 2:
        if all(ws.cell(row=2, column=col).value in (None, "") for col in range(1, ws.max_column + 1)):
            return 2
    return ws.max_row + 1


def _format_runs_row(ws, row_idx: int) -> None:
    ws.cell(row=row_idx, column=5).number_format = "yyyy-mm-dd"
    ws.cell(row=row_idx, column=6).number_format = "hh:mm:ss"
    ws.cell(row=row_idx, column=7).number_format = "hh:mm:ss"


def _add_success_line_chart(ws, anchor: str) -> None:
    chart = LineChart()
    chart.height = 8
    chart.width = 14
    chart.y_axis.title = "Taxa"
    chart.y_axis.number_format = "0%"
    chart.title = "Taxa de Sucesso"

    data = Reference(ws.parent[RUNS_SHEET], min_col=12, min_row=1, max_row=1000)
    cats = Reference(ws.parent[RUNS_SHEET], min_col=5, min_row=2, max_row=1000)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def _add_volume_bar_chart(ws, anchor: str) -> None:
    chart = BarChart()
    chart.height = 8
    chart.width = 14
    chart.title = "Processado vs Erros"
    chart.y_axis.title = "Volume"
    data_proc = Reference(ws.parent[RUNS_SHEET], min_col=9, max_col=9, min_row=1, max_row=1000)
    data_err = Reference(ws.parent[RUNS_SHEET], min_col=11, max_col=11, min_row=1, max_row=1000)
    cats = Reference(ws.parent[RUNS_SHEET], min_col=5, min_row=2, max_row=1000)
    chart.add_data(data_proc, titles_from_data=True)
    chart.add_data(data_err, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def _add_error_pie_chart(ws, anchor: str) -> None:
    ws["L3"] = "Tecnico"
    ws["L4"] = "Regra de Negocio"
    ws["L5"] = "Dados Invalidos"
    ws["L6"] = "Externo/Indisponibilidade"
    ws["M2"] = "Qtd"
    ws["M3"] = '=COUNTIF(tbl_errors[tipo_erro],"Tecnico")'
    ws["M4"] = '=COUNTIF(tbl_errors[tipo_erro],"Regra de Negocio")'
    ws["M5"] = '=COUNTIF(tbl_errors[tipo_erro],"Dados Invalidos")'
    ws["M6"] = '=COUNTIF(tbl_errors[tipo_erro],"Externo/Indisponibilidade")'

    chart = PieChart()
    chart.height = 8
    chart.width = 10
    chart.title = "Distribuicao de Erros"
    data = Reference(ws, min_col=13, min_row=2, max_row=6)
    cats = Reference(ws, min_col=12, min_row=3, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def _add_duration_line_chart(ws, anchor: str) -> None:
    chart = LineChart()
    chart.height = 8
    chart.width = 14
    chart.title = "Duracao por Execucao"
    chart.y_axis.title = "Segundos"
    data = Reference(ws.parent[RUNS_SHEET], min_col=8, min_row=1, max_row=1000)
    cats = Reference(ws.parent[RUNS_SHEET], min_col=5, min_row=2, max_row=1000)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def _calc_resultado_final(
    total_processado,
    total_sucesso,
    total_erro,
    erros_auto,
    erros_manuais,
) -> str:
    total_processado = _safe_int(total_processado) or 0
    total_sucesso = _safe_int(total_sucesso) or 0
    total_erro = _safe_int(total_erro) or 0
    erros_auto = _safe_int(erros_auto) or 0
    erros_manuais = _safe_int(erros_manuais) or 0

    if total_processado == 0:
        return "Incompleto"

    pendentes = max(total_erro - erros_auto - erros_manuais, 0)
    if total_sucesso == total_processado and pendentes == 0:
        return "100% Concluído"
    if total_sucesso + total_erro != total_processado:
        return "Incompleto"
    return "Concluído com Exceções"


def _safe_int(value) -> Optional[int]:
    try:
        if value is None:
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_dt(value) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue
    return None


def _save_pending(wb: Workbook, run_data: Dict, errors: List[Dict]) -> None:
    os.makedirs(PENDING_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pending_xlsx = os.path.join(PENDING_DIR, f"Auditoria_Robos__PENDENTE__{timestamp}.xlsx")
    pending_json = os.path.join(PENDING_DIR, f"run__{run_data.get('run_id')}.json")

    try:
        wb.save(pending_xlsx)
        logger.warning("Arquivo principal em uso. Salvando pendente em %s", pending_xlsx)
    except Exception as exc:
        logger.error("Falha ao salvar Excel pendente: %s", exc)

    try:
        payload = {"run_data": _json_safe(run_data), "errors": _json_safe(errors)}
        with open(pending_json, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        logger.error("Falha ao salvar JSON pendente: %s", exc)


def _json_safe(value):
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def get_project_version() -> str:
    candidates = [
        os.path.join(os.getcwd(), "version.txt"),
        os.path.join(os.path.dirname(__file__), "..", "version.txt"),
        os.path.join(os.path.dirname(__file__), "version.txt"),
    ]
    for path in candidates:
        try:
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    return f.read().strip()
        except Exception:
            continue
    return "1.0.0"


def _try_get_git_commit() -> str:
    head_path = os.path.join(os.getcwd(), ".git", "HEAD")
    try:
        if not os.path.exists(head_path):
            return ""
        with open(head_path, "r", encoding="utf-8") as f:
            ref = f.read().strip()
        if ref.startswith("ref:"):
            ref_path = ref.split(" ", 1)[1].strip()
            ref_file = os.path.join(os.getcwd(), ".git", ref_path.replace("/", os.sep))
            with open(ref_file, "r", encoding="utf-8") as f:
                commit = f.read().strip()
        else:
            commit = ref
        return commit[:8]
    except Exception:
        return ""
